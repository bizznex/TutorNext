from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "product vision" / "ContexureEarth_Tutor_Professional_Project_Planner.xlsx"


STATUS = ["Done", "In Progress", "In Review", "In QA", "Blocked", "Not Started", "Deferred"]
PRIORITY = ["P0", "P1", "P2", "P3"]
PHASES = [
    ("D0", "Already Done / Existing BizzNexx Assets", date(2026, 6, 1), date(2026, 6, 13)),
    ("P0", "Foundation & Product Definition", date(2026, 6, 15), date(2026, 6, 28)),
    ("P1", "Concierge Prototype", date(2026, 6, 29), date(2026, 7, 19)),
    ("P2", "Private Beta", date(2026, 7, 20), date(2026, 9, 6)),
    ("P3", "Paid Pilot", date(2026, 9, 7), date(2026, 12, 13)),
    ("P4", "Public Launch V1", date(2026, 12, 14), date(2027, 3, 14)),
    ("P5", "Growth & Intelligence V1", date(2027, 3, 15), date(2027, 9, 12)),
    ("P6", "Learning Context V1", date(2027, 9, 13), date(2028, 3, 12)),
    ("P7", "Institution Expansion", date(2028, 3, 13), date(2029, 6, 30)),
]

WORKSTREAMS = {
    "STR": "Strategy & Product Definition",
    "RES": "Research & Customer Discovery",
    "UX": "UX, Design & Copy",
    "FE": "Frontend Engineering",
    "BE": "Backend & Data Model",
    "SEC": "Auth, Privacy & Trust",
    "WA": "WhatsApp & Messaging",
    "DATA": "Analytics & Metrics",
    "OPS": "Operator Setup & Support",
    "GTM": "Marketing, Sales & Launch",
    "QA": "QA, Testing & Release Gates",
    "INFRA": "Infrastructure & Deployment",
    "AI": "AI & Learning Intelligence",
    "INST": "Institution Expansion",
    "DOC": "Documentation & Handoff",
}


def add_days(d: date, days: int) -> date:
    return d + timedelta(days=days)


def task(
    task_id,
    phase,
    ws,
    title,
    description,
    owner,
    priority,
    status,
    start,
    end,
    deps,
    acceptance,
    validation,
    vision,
    effort=3,
):
    return [
        task_id,
        phase,
        WORKSTREAMS[ws],
        title,
        description,
        owner,
        priority,
        status,
        start,
        end,
        deps,
        effort,
        acceptance,
        validation,
        vision,
        "",
    ]


DONE_TASKS = [
    task("DONE-001", "D0", "STR", "Learning Studio strategy created", "Created Learning Studio business ontology, strategy plan, execution plan, and beta rollout readiness plan.", "Founder/Product", "P0", "Done", date(2026, 6, 1), date(2026, 6, 6), "", "Strategy docs exist and define tutor wedge.", "Learning Studio is defined as trust, trial, parent communication, and recurring-fee business.", "V1/V2/V4/V5"),
    task("DONE-002", "D0", "STR", "First validation wedge narrowed", "Narrowed initial beta wedge to Ananya Learning Studio and tutor/coach persona.", "Founder/Product", "P0", "Done", date(2026, 6, 5), date(2026, 6, 6), "DONE-001", "Ananya is the demo anchor.", "Product can be reviewed through one concrete tutor story.", "V1"),
    task("DONE-003", "D0", "FE", "Parallel Beta Studio UI added", "Added Beta Studio tab for tutor businesses while preserving existing prototype screens.", "Frontend", "P0", "Done", date(2026, 6, 6), date(2026, 6, 7), "DONE-002", "Tutor-specific surface exists.", "Reviewer can open Ananya Learning Studio Beta Studio.", "V1/V7"),
    task("DONE-004", "D0", "FE", "Parent-ready Learning Studio surface built", "Added teacher credibility, proof, course cards, trial inquiry form, parent journey, trial action queue, monthly fee actions, and FAQ.", "Frontend/Product", "P0", "Done", date(2026, 6, 6), date(2026, 6, 7), "DONE-003", "Core tutor beta UI exists.", "The demo can show parent page to fee workflow.", "V1/V2/V4/V5/V6"),
    task("DONE-005", "D0", "SEC", "Beta safety notes added", "Added beta disclaimer, beta terms view, and inquiry consent/data-use copy.", "Security/Product", "P0", "Done", date(2026, 6, 7), date(2026, 6, 8), "DONE-004", "Beta users see scope and data-use notes.", "No real-data review without safety copy.", "V7"),
    task("DONE-006", "D0", "BE", "Ananya seed data enriched", "Aligned backend demo seed data with richer Learning Studio story, persona statuses, and message templates.", "Backend/Product", "P0", "Done", date(2026, 6, 7), date(2026, 6, 8), "DONE-004", "Seed data supports realistic demo.", "Ananya data includes courses, proof, FAQs, leads, payments.", "V1/V2/V5"),
    task("DONE-007", "D0", "SEC", "Private beta access controls added", "Added optional frontend access code, backend API key gate, environment examples, and production-safe settings hooks.", "Security/Backend/Frontend", "P0", "Done", date(2026, 6, 8), date(2026, 6, 9), "DONE-005", "Beta access can be restricted.", "Demo can be shared with selected reviewers only.", "V7"),
    task("DONE-008", "D0", "INFRA", "Deployment blueprint prepared", "Added Postgres support, Gunicorn, static root, render.yaml, health endpoint, and deployment docs.", "Engineering", "P0", "Done", date(2026, 6, 9), date(2026, 6, 10), "DONE-007", "Render deployment path exists.", "Backend health endpoint available.", "V7"),
    task("DONE-009", "D0", "INFRA", "GitHub Actions CI added", "Added frontend build and backend checks on main.", "Engineering/QA", "P0", "Done", date(2026, 6, 10), date(2026, 6, 10), "DONE-008", "CI workflow exists.", "Build/check pipeline can run on main.", "V7"),
    task("DONE-010", "D0", "INFRA", "GitHub Pages preview workflow added", "Added no-new-cloud-account frontend preview path for UI-only guided review.", "Engineering", "P1", "Done", date(2026, 6, 10), date(2026, 6, 10), "DONE-009", "Preview workflow and guide exist.", "UI can be reviewed without backend hosting.", "V1"),
    task("DONE-011", "D0", "DOC", "Root README updated", "Updated README with BizzNexx product shape, Learning Studio beta, run commands, and deployment paths.", "Documentation", "P0", "Done", date(2026, 6, 10), date(2026, 6, 11), "DONE-008", "README gives clear project entry.", "New reviewer can understand current status.", "V7"),
    task("DONE-012", "D0", "OPS", "Beta review guide created", "Created guided review script for tutor beta reviews.", "Operations/Product", "P0", "Done", date(2026, 6, 11), date(2026, 6, 11), "DONE-004", "Review guide covers parent page, inquiry, action queue, fees, and WTP.", "Beta calls have a consistent script.", "V1/V2/V5"),
    task("DONE-013", "D0", "OPS", "Beta feedback form created", "Created structured beta feedback form with workflow fit, commercial signal, trust/safety, and product gaps.", "Operations/Product", "P0", "Done", date(2026, 6, 11), date(2026, 6, 11), "DONE-012", "Feedback form can be used per reviewer.", "WTP and product-fit signals are captured.", "V7"),
    task("DONE-014", "D0", "QA", "Private beta QA checklist created", "Created hosted deployment QA checks before sharing beta URL.", "QA", "P0", "Done", date(2026, 6, 11), date(2026, 6, 12), "DONE-008", "Checklist covers HTTPS, access gate, Ananya, inquiry, message, payment, beta terms.", "No beta sharing without QA pass.", "V7"),
    task("DONE-015", "D0", "STR", "Tutor growth feature universe created", "Created feature strategy, persona mapping, heatmap, ranking model, and top build order.", "Product", "P0", "Done", date(2026, 6, 12), date(2026, 6, 12), "DONE-001", "Feature universe and ranking are documented.", "MVP spine is clear.", "V1-V8"),
    task("DONE-016", "D0", "STR", "Production accountable tasklist created", "Created accountable workstreams across product, UX, FE, BE, auth, analytics, QA, infra, ops, support, GTM, and risk.", "Product", "P0", "Done", date(2026, 6, 12), date(2026, 6, 13), "DONE-015", "Production task board exists.", "Team can execute using task IDs.", "V7"),
    task("DONE-017", "D0", "DOC", "Team handoff guide created", "Created TEAM_HANDOFF_START_HERE.md with mission, decisions, MVP spine, first tasks, repo structure, and QA gates.", "Documentation/Product", "P0", "Done", date(2026, 6, 13), date(2026, 6, 13), "DONE-016", "Team entry point exists.", "New contributor can start from handoff guide.", "V7"),
    task("DONE-018", "D0", "STR", "ContexureEarth vision mind map created", "Created product vision mind map and renamed vision file under ContexureEarth.", "Founder/Product", "P1", "Done", date(2026, 6, 13), date(2026, 6, 13), "", "Vision file exists under final name.", "Long-term learning intelligence direction captured.", "V8/V9/V10"),
    task("DONE-019", "D0", "STR", "ContexureEarth tutor roadmap created", "Created research-backed tutor product roadmap linking BizzNexx work to ContexureEarth.", "Product", "P1", "Done", date(2026, 6, 13), date(2026, 6, 13), "DONE-018", "Roadmap covers research, build, GTM, launch, post-launch.", "Execution path is visible.", "V1-V10"),
    task("DONE-020", "D0", "STR", "BizzNexx critical evaluation created", "Evaluated what BizzNexx already achieves and how it should evolve into ContexureEarth Tutor.", "Product", "P1", "Done", date(2026, 6, 13), date(2026, 6, 13), "DONE-019", "Evaluation gives brand architecture and product migration recommendation.", "Decision: ContexureEarth public brand; BizzNexx operating engine.", "V1-V10"),
    task("DONE-021", "D0", "QA", "Initial verification passed", "npm run build, Django check, collectstatic dry run, compileall, and local UI response were verified previously.", "Engineering/QA", "P0", "Done", date(2026, 6, 10), date(2026, 6, 13), "DONE-009", "Verification evidence recorded in project log.", "Prototype is stable enough for deployment prep.", "V7"),
    task("DONE-022", "D0", "GTM", "Code pushed to GitHub", "Project log records code pushed with key commits for strategy, UI, safety, deployment, and preview.", "Engineering", "P0", "Done", date(2026, 6, 10), date(2026, 6, 13), "DONE-021", "Commits are recorded.", "Remote repo has current work.", "V7"),
]


FUTURE_TASKS = [
    # immediate deployment and beta gap
    task("IMD-001", "P0", "INFRA", "Create Render Blueprint deployment", "Create Render Blueprint from GitHub and let it provision frontend, backend, and Postgres database.", "Engineering", "P0", "Not Started", date(2026, 6, 15), date(2026, 6, 17), "DONE-008,DONE-022", "Render services exist and deploy successfully.", "Hosted API and frontend URLs available.", "V7", 4),
    task("IMD-002", "P0", "SEC", "Wire beta API keys", "Copy backend BIZZNEXX_BETA_API_KEY to frontend VITE_BETA_API_KEY and set VITE_BETA_ACCESS_CODE.", "Security/Engineering", "P0", "Not Started", date(2026, 6, 17), date(2026, 6, 18), "IMD-001,DONE-007", "Frontend and backend keys match; access gate works.", "Unauthorized beta access blocked.", "V7", 2),
    task("IMD-003", "P0", "QA", "Run hosted beta QA checklist", "Confirm HTTPS, access gate, Ananya, Beta Studio, inquiry submit, lead creation, message preview, mark paid, beta terms.", "QA/Product", "P0", "Not Started", date(2026, 6, 18), date(2026, 6, 20), "IMD-002,DONE-014", "QA checklist completed with pass/fail evidence.", "No beta user sees broken hosted flow.", "V7", 3),
    task("IMD-004", "P0", "RES", "Recruit first 5 tutor reviewers", "Use existing networks to schedule guided review calls with academic tutors/coaches.", "Founder/Research", "P0", "Not Started", date(2026, 6, 18), date(2026, 6, 24), "DONE-012,DONE-013", "5 calls scheduled or completed.", "At least 3 reviewers fit target ICP.", "V1/V2/V5", 3),
    task("IMD-005", "P0", "RES", "Capture beta feedback in structured form", "Use BETA_FEEDBACK_FORM for every reviewer and summarize share-worthiness, WTP, trust, and gaps.", "Research/Product", "P0", "Not Started", date(2026, 6, 20), date(2026, 6, 28), "IMD-003,IMD-004", "Each review has completed feedback form.", "Evidence informs next sprint.", "V7", 3),
    task("IMD-006", "P0", "STR", "Founder decision: brand architecture", "Confirm whether public brand is ContexureEarth Tutor with BizzNexx as internal/legacy engine.", "Founder", "P0", "Not Started", date(2026, 6, 20), date(2026, 6, 22), "DONE-020", "Decision recorded in roadmap and handoff docs.", "No confusion in UI/demo language.", "V1-V10", 2),
    task("IMD-007", "P0", "UX", "UI copy pass for ContexureEarth Tutor", "Replace generic microbusiness language in tutor path with learner, parent, course, batch, fee cycle, operating rhythm.", "UX/Product", "P0", "Not Started", date(2026, 6, 22), date(2026, 6, 27), "IMD-006,FE-001", "Tutor demo uses education-native language.", "Reviewers do not perceive it as generic CRM.", "V1/V8", 4),
    # foundation
    task("C-001", "P0", "FE", "Audit current frontend architecture", "Identify routes, state, API usage, design debt, responsive gaps, and what can be reused.", "Frontend", "P0", "Not Started", date(2026, 6, 15), date(2026, 6, 18), "DONE-004", "Frontend audit is documented.", "Large implementation changes have a clear base.", "V7", 4),
    task("D-001", "P0", "BE", "Audit backend models/APIs", "Identify model gaps, workflow gaps, auth gaps, migration risks, and generic-to-education naming pressure.", "Backend", "P0", "Not Started", date(2026, 6, 15), date(2026, 6, 18), "DONE-006", "Backend audit is documented.", "Tenant/auth/model decisions can be made.", "V7/V8", 4),
    task("A-001", "P0", "STR", "Define production personas", "Document academic tutor, exam coach, hobby teacher, skill trainer, fitness trainer, corporate trainer, and priority order.", "Product", "P0", "Not Started", date(2026, 6, 15), date(2026, 6, 17), "DONE-015", "Persona definitions are approved.", "P-AST remains first ICP unless evidence changes.", "V1", 2),
    task("A-004", "P0", "STR", "Map public visitor journey", "Map parent path from shared link to inquiry, trial CTA, confirmation, and WhatsApp handoff.", "Product/UX", "P0", "Not Started", date(2026, 6, 17), date(2026, 6, 20), "A-001", "Journey includes trust, course fit, fee clarity, trial, and privacy.", "Parent can understand next step in 30 seconds.", "V1/V2", 3),
    task("A-005", "P0", "STR", "Map owner inquiry-to-fee journey", "Map owner path from inquiry to trial to enrollment to fee tracking and proof/referral prompt.", "Product/UX", "P0", "Not Started", date(2026, 6, 17), date(2026, 6, 21), "A-001", "Every owner step has status, next action, and message template.", "Owner knows what to do today.", "V2/V4/V5/V6/V7", 3),
    task("D-002", "P0", "BE", "Define tenant/account model", "Decide account, owner, operator, support, tutor studio boundaries and data isolation rules.", "Backend/Security", "P0", "Not Started", date(2026, 6, 19), date(2026, 6, 24), "D-001", "Multi-tutor data isolation is designed.", "No real beta without tenant boundary.", "V7", 5),
    task("E-001", "P0", "SEC", "Choose production auth approach", "Choose email/password, OTP, managed auth, or phased approach for owner/operator dashboard.", "Security/Backend", "P0", "Not Started", date(2026, 6, 22), date(2026, 6, 26), "D-002", "Auth decision covers roles, recovery, beta path, and production migration.", "Owner dashboard protected.", "V7", 4),
    task("N-001", "P0", "INFRA", "Choose production hosting architecture", "Decide staging/production deployment, database, static assets, backups, monitoring, domain, rollback.", "Engineering", "P0", "Not Started", date(2026, 6, 22), date(2026, 6, 27), "D-001,E-001", "Hosting architecture is approved.", "Staging can be created confidently.", "V7", 4),
    task("B-001", "P0", "UX", "Define ContexureEarth Tutor visual direction", "Set calm, credible, parent-friendly, mobile-first, education-native design principles.", "UX", "P0", "Not Started", date(2026, 6, 19), date(2026, 6, 23), "IMD-006", "Moodboard/design principles approved.", "Tutor would not be embarrassed to share page.", "V1", 3),
    task("B-002", "P0", "UX", "Responsive layout system", "Define layout system for public page, owner dashboard, and operator console across mobile/desktop widths.", "UX/Frontend", "P0", "Not Started", date(2026, 6, 24), date(2026, 7, 1), "B-001,C-001", "Layouts work at 360, 390, 768, 1024, 1440 widths.", "No text overlap or awkward wrapping.", "V1/V7", 5),
    task("D-004", "P0", "BE", "Define Course/Batch model", "Define course, batch, schedule, capacity, fee policy, board/class/subject fields.", "Backend/Product", "P0", "Not Started", date(2026, 6, 25), date(2026, 7, 2), "D-002,A-005", "Model supports academic tutor MVP and future exam/hobby variants.", "Course cards can be API-backed.", "V1/V2", 5),
    task("D-005", "P0", "BE", "Define Inquiry model", "Define structured parent inquiry with academic context, consent, source, duplicate detection path.", "Backend/Product", "P0", "Not Started", date(2026, 6, 25), date(2026, 7, 2), "D-002,A-004", "Inquiry stores class, board, subject, concern, timing, trial interest, source.", "Owner avoids repeated basic questions.", "V2", 5),
    task("D-006", "P0", "BE", "Define Lead/Trial states", "Define parent inquiry, contacted, trial proposed, scheduled, attended, no-show, enrolled, lost, follow-later states.", "Backend/Product", "P0", "Not Started", date(2026, 6, 28), date(2026, 7, 5), "D-005,A-005", "States cover inquiry-to-enrollment journey.", "Trial outcomes are measurable.", "V4", 5),
    task("F-003", "P0", "WA", "Create WhatsApp template library", "Create first response, trial options, trial reminder, post-trial, fee reminder, progress, proof, referral messages.", "Product/Copy", "P0", "Not Started", date(2026, 6, 28), date(2026, 7, 4), "A-005,IMD-005", "Templates are polite, editable, parent-safe, context-aware.", "Tutors say tone feels natural.", "V3/V4/V5/V6", 4),
    task("C-003", "P0", "FE", "Implement public studio page route", "Productionize share-ready public page from API-backed tutor profile, courses, proof, FAQ, and CTA.", "Frontend", "P0", "Not Started", date(2026, 7, 1), date(2026, 7, 12), "B-002,D-004", "Page renders loading, error, fallback, and published states.", "Tutor says they would share page.", "V1", 7),
    task("C-004", "P0", "FE", "Implement tutor inquiry flow", "Create mobile-friendly inquiry form and confirmation state with consent and source tracking.", "Frontend", "P0", "Not Started", date(2026, 7, 8), date(2026, 7, 18), "C-003,D-005", "Inquiry submits successfully and creates backend lead.", "Parent knows what happens next.", "V2", 6),
    task("F-001", "P0", "WA", "Implement manual WhatsApp click-to-chat", "Open WhatsApp with prefilled context from inquiry/trial/fee action and copy fallback.", "Frontend", "P0", "Not Started", date(2026, 7, 12), date(2026, 7, 20), "C-004,F-003", "Owner can open/edit/copy WhatsApp messages.", "Manual handoff works before API automation.", "V3", 5),
    task("C-005", "P0", "FE", "Implement owner Today dashboard", "Create action queue for new inquiries, trials, fee dues, overdue payments, proof/referral prompts.", "Frontend", "P0", "Not Started", date(2026, 7, 14), date(2026, 7, 25), "A-005,D-006,F-001", "Owner sees actionable cards and can update statuses.", "Owner knows what needs attention in under 30 seconds.", "V7", 7),
    task("C-006", "P0", "FE", "Implement leads and trial workflow UI", "Owner can move inquiry through contacted, trial, enrolled, lost, follow-later states.", "Frontend", "P0", "Not Started", date(2026, 7, 22), date(2026, 8, 2), "C-005,D-006", "Owner can update trial statuses without confusion.", "Trial queue has no stale invisible leads.", "V4", 6),
    task("C-007", "P0", "FE", "Implement student/learner memory UI", "Show lightweight learner profile and notes connected to inquiry, enrollment, parent contact, next action.", "Frontend", "P0", "Not Started", date(2026, 7, 28), date(2026, 8, 10), "C-006,D-004,D-005", "Owner can view learner context and next action.", "Tutor sees value beyond CRM.", "V8", 7),
    task("C-008", "P0", "FE", "Implement fee tracker UI", "Owner can see due/overdue/part-paid/paid, preview reminders, and mark paid.", "Frontend", "P0", "Not Started", date(2026, 8, 1), date(2026, 8, 12), "C-007,D-004", "Owner identifies fee dues and sends reminder.", "At least one real fee cycle tracked.", "V5", 6),
    task("E-006", "P0", "SEC", "Add inquiry consent timestamp", "Store consent version/timestamp with every public inquiry and display data-use note.", "Security/Backend/Frontend", "P0", "Not Started", date(2026, 7, 8), date(2026, 7, 18), "D-005,C-004", "Every inquiry has consent metadata.", "No real data without consent trail.", "V7", 4),
    task("M-001", "P0", "QA", "Inquiry-to-lead backend tests", "Create backend tests for public inquiry capture, lead creation, and follow-up task creation.", "QA/Backend", "P0", "Not Started", date(2026, 7, 18), date(2026, 7, 28), "D-005,D-006,C-004", "Tests pass in CI.", "Workflow cannot silently break.", "V2/V7", 5),
    task("N-002", "P0", "INFRA", "Create staging environment", "Create staging deployment mirroring production configuration and seeded demo data.", "Engineering", "P0", "Not Started", date(2026, 7, 20), date(2026, 8, 2), "N-001,E-001,M-001", "Staging is stable and separate from production.", "Guided reviews use stable URL.", "V7", 5),
    # private beta production-grade gaps
    task("SEC-010", "P0", "SEC", "Implement real owner login", "Replace access-code-only beta gating with real owner authentication for dashboard.", "Security/Backend/Frontend", "P0", "Not Started", date(2026, 8, 1), date(2026, 8, 18), "E-001,N-002", "Owner dashboard requires authenticated owner.", "Private user data is not public.", "V7", 8),
    task("SEC-011", "P0", "SEC", "Implement operator/admin login", "Protect setup and operator console with role-specific authentication.", "Security/Backend/Frontend", "P0", "Not Started", date(2026, 8, 8), date(2026, 8, 22), "SEC-010", "Operator tools are not publicly accessible.", "No external user can access setup console.", "V7", 6),
    task("SEC-012", "P0", "SEC", "Role-based access control", "Implement roles for owner, operator, admin, support with least privilege.", "Security/Backend", "P0", "Not Started", date(2026, 8, 15), date(2026, 8, 30), "SEC-011,D-002", "Permissions are enforced across APIs.", "Cross-role tests pass.", "V7", 7),
    task("SEC-013", "P0", "SEC", "Privacy policy and beta terms review", "Prepare privacy policy, beta terms, data deletion promise, no guarantee claims, minor data caution.", "Security/Product", "P0", "Not Started", date(2026, 8, 15), date(2026, 8, 31), "E-006,IMD-005", "Docs are reviewed and linked from beta.", "Users understand data and scope.", "V7", 4),
    task("SEC-014", "P0", "SEC", "Data deletion workflow", "Create manual/admin deletion or anonymization process for parent/student records.", "Security/Backend/OPS", "P0", "Not Started", date(2026, 8, 22), date(2026, 9, 4), "SEC-012,SEC-013", "Deletion request can be fulfilled and logged.", "User data control exists.", "V7", 5),
    task("INFRA-010", "P0", "INFRA", "Configure backups and restore test", "Automate database backups and run at least one restore test.", "Engineering/Security", "P0", "Not Started", date(2026, 8, 22), date(2026, 9, 4), "N-002", "Backup policy and restore evidence exist.", "No real beta without recoverability.", "V7", 5),
    task("INFRA-011", "P0", "INFRA", "Configure logs and error monitoring", "Capture backend/frontend errors and uptime alerts for beta.", "Engineering", "P0", "Not Started", date(2026, 8, 22), date(2026, 9, 4), "N-002", "Errors are visible and triageable.", "Support can respond to issues.", "V7", 4),
    task("QA-010", "P0", "QA", "Responsive visual QA matrix", "Check core screens at mobile and desktop widths and record defects.", "QA/UX", "P0", "Not Started", date(2026, 8, 20), date(2026, 9, 6), "C-003,C-004,C-005,C-008", "Screens pass 360, 390, 768, 1024, 1440 widths.", "No text overlap or broken mobile layout.", "V1/V7", 5),
    task("QA-011", "P0", "QA", "Private beta release gate review", "Run full private beta blocker checklist before inviting real users.", "QA/Product", "P0", "Not Started", date(2026, 9, 1), date(2026, 9, 6), "SEC-014,INFRA-010,INFRA-011,QA-010", "Gate passes or blockers are documented.", "Real users are not invited prematurely.", "V7", 3),
    # paid pilot and launch
    task("GTM-001", "P1", "GTM", "Define Launch Pack offer", "Define setup scope, price, deliverables, exclusions, and support level.", "Founder/GTM", "P0", "Not Started", date(2026, 9, 7), date(2026, 9, 14), "IMD-005,QA-011", "Offer is simple and sellable.", "Tutors can understand why to pay.", "V1/V7", 3),
    task("GTM-002", "P1", "GTM", "Define Monthly Assist Pack", "Define monthly value around updates, dashboard, fee/follow-up support, proof/referral prompts.", "Founder/GTM", "P1", "Not Started", date(2026, 9, 10), date(2026, 9, 18), "GTM-001", "Monthly pack has clear recurring value.", "Tutor can explain why subscription continues.", "V7", 3),
    task("GTM-003", "P1", "GTM", "Create sales demo script", "Create 12-minute demo showing page, inquiry, trial, WhatsApp, fee, proof, dashboard.", "Founder/GTM", "P0", "Not Started", date(2026, 9, 15), date(2026, 9, 24), "GTM-001,C-005,C-008", "Demo script is repeatable.", "Prospect understands product without roadmap lecture.", "V1/V7", 3),
    task("OPS-001", "P1", "OPS", "Operator setup console V1", "Implement setup checklist, readiness score, blockers, missing proof, template readiness.", "OPS/Frontend/Backend", "P1", "Not Started", date(2026, 9, 15), date(2026, 10, 10), "C-003,C-004,F-003", "Operator can set up second tutor in under 90 minutes.", "Assisted setup becomes repeatable.", "V1/V7", 8),
    task("DATA-001", "P1", "DATA", "Source tracking dashboard", "Show which source creates inquiries and which converts to trial/enrollment.", "Data/Frontend", "P1", "Not Started", date(2026, 9, 22), date(2026, 10, 15), "DATA-000,C-006", "Owner sees source-wise lead quality.", "Tutor changes marketing behavior.", "V7", 6),
    task("DATA-000", "P0", "DATA", "Define event taxonomy", "Define page view, CTA click, inquiry, trial, fee, message, proof, setup, owner action events.", "Data/Product", "P0", "Not Started", date(2026, 7, 8), date(2026, 7, 18), "A-004,A-005", "Events have trigger, properties, privacy level, owner.", "Activation funnel can be measured.", "V7", 4),
    task("DATA-002", "P1", "DATA", "Weekly owner digest", "Create weekly digest of inquiries, trials, fees, proof prompts, and next actions.", "Data/Backend/Frontend", "P1", "Not Started", date(2026, 10, 1), date(2026, 10, 25), "DATA-000,C-005,C-008", "Digest is actionable and concise.", "Tutors check weekly.", "V7", 6),
    task("FE-020", "P1", "FE", "Testimonial request workflow", "Request, review, approve, hide, and publish testimonials/improvement stories.", "Frontend/Backend", "P1", "Not Started", date(2026, 10, 8), date(2026, 11, 1), "C-003,F-003", "Proof request loop exists.", "Tutor collects new proof.", "V6", 6),
    task("FE-021", "P1", "FE", "Referral request workflow", "Request referrals after positive milestone, result, trial, or monthly progress.", "Frontend/Backend", "P1", "Not Started", date(2026, 10, 15), date(2026, 11, 8), "FE-020,F-003", "Referral prompts are available and trackable.", "First referral request sent.", "V6", 5),
    task("GTM-004", "P1", "GTM", "Close first 10 paid pilots", "Convert beta and warm prospects into paid pilot accounts.", "Founder/GTM", "P0", "Not Started", date(2026, 11, 1), date(2026, 12, 13), "GTM-003,OPS-001,DATA-002", "10 paid pilots or equivalent signed commitments.", "Revenue signal validated.", "V7", 8),
    task("GTM-005", "P1", "GTM", "Create 3 case studies", "Create before/after workflow proof from paid pilots.", "Founder/GTM", "P1", "Not Started", date(2026, 11, 15), date(2026, 12, 31), "GTM-004,DATA-002", "Case studies include metrics, screenshots, quote, problem, outcome.", "Proof supports public launch.", "V6/V7", 5),
    task("INFRA-020", "P1", "INFRA", "Production environment", "Create production environment with domain, HTTPS, database, backups, monitoring, rollback.", "Engineering", "P1", "Not Started", date(2026, 12, 14), date(2027, 1, 15), "GTM-004,INFRA-010,INFRA-011", "Production deployment is stable and recoverable.", "Public launch can proceed.", "V7", 8),
    task("FE-030", "P1", "FE", "Self-serve onboarding V1", "Allow tutor to create/complete profile with assisted fallback.", "Frontend/Backend", "P1", "Not Started", date(2027, 1, 1), date(2027, 1, 31), "OPS-001,SEC-010", "Tutor reaches first value without full manual setup.", "First value under 30 minutes.", "V1/V7", 8),
    task("GTM-010", "P1", "GTM", "Public landing and pricing page", "Publish ContexureEarth Tutor landing page, pricing, demo, FAQ, case studies, privacy.", "Founder/GTM/Frontend", "P1", "Not Started", date(2027, 1, 15), date(2027, 2, 10), "GTM-005,INFRA-020", "Landing page explains offer and converts demo interest.", "Demo requests begin.", "V1/V7", 5),
    task("OPS-010", "P1", "OPS", "Support help center", "Create user docs for setup, inquiry, WhatsApp, fees, proof, privacy, troubleshooting.", "Support/OPS", "P1", "Not Started", date(2027, 1, 20), date(2027, 2, 15), "OPS-001", "Help center covers common tasks.", "Support time per account reduces.", "V7", 4),
    task("DATA-010", "P1", "DATA", "Activation and retention dashboard", "Track setup, share, inquiry, action completion, fee usage, retention, churn, support burden.", "Data/Product", "P1", "Not Started", date(2027, 2, 1), date(2027, 2, 28), "DATA-002,INFRA-020", "Founder can review activation weekly.", "Roadmap decisions use real data.", "V7", 6),
    task("QA-020", "P1", "QA", "Public launch QA gate", "Run production monitoring, privacy, deletion, onboarding, billing, conversion, mobile, rollback checks.", "QA/Product", "P1", "Not Started", date(2027, 2, 15), date(2027, 3, 7), "INFRA-020,FE-030,GTM-010,DATA-010", "Public launch checklist passes.", "No P0 blocker before campaign.", "V7", 5),
    task("GTM-011", "P1", "GTM", "Controlled public launch campaign", "Launch to academic tutors and small coaching centers with content, demos, referrals, case studies.", "Founder/GTM", "P1", "Not Started", date(2027, 3, 1), date(2027, 3, 14), "QA-020,GTM-010", "Campaign assets go live.", "50 paying customers or strong pipeline.", "V1/V7", 6),
    # intelligence and learning future
    task("FE-040", "P2", "FE", "Batch capacity and waitlist", "Add batch seats, filling/full/waitlist states, and trial-fit routing.", "Frontend/Backend", "P2", "Not Started", date(2027, 3, 15), date(2027, 4, 15), "GTM-011", "Batch availability informs owner actions.", "Tutors avoid overbooking.", "V4/V7", 6),
    task("FE-041", "P2", "FE", "Attendance-lite", "Track class attendance and trigger absence/makeup prompts.", "Frontend/Backend", "P2", "Not Started", date(2027, 4, 1), date(2027, 5, 1), "FE-040", "Attendance can be marked quickly by batch.", "Absence follow-ups are sent.", "V8", 6),
    task("FE-042", "P2", "FE", "Parent progress update templates", "Create monthly progress update with covered topics, strengths, concerns, next focus.", "Frontend/Product", "P2", "Not Started", date(2027, 5, 1), date(2027, 6, 15), "FE-041,F-003", "Tutor can send progress update without heavy data entry.", "Progress updates improve retention.", "V8", 6),
    task("AI-001", "P2", "AI", "AI inquiry summary", "Summarize inquiry into parent concern, class/board/subject, urgency, and suggested next message.", "AI/Product", "P2", "Not Started", date(2027, 4, 15), date(2027, 5, 31), "DATA-010,C-004", "AI summary is editable and never auto-sends.", "30%+ summaries accepted/edited.", "V2/V8", 6),
    task("AI-002", "P2", "AI", "AI message drafts", "Draft parent replies, fee reminders, progress notes, testimonial and referral messages with tone controls.", "AI/Product", "P2", "Not Started", date(2027, 5, 15), date(2027, 7, 1), "AI-001,F-003", "AI drafts are safe, editable, and context-aware.", "Message writing time reduces.", "V3/V5/V8", 6),
    task("DATA-020", "P2", "DATA", "Source quality dashboard", "Rank inquiry sources by volume, conversion, quality, and lost reasons.", "Data/Frontend", "P2", "Not Started", date(2027, 6, 1), date(2027, 7, 15), "DATA-010,DATA-001", "Owner sees which channels create good inquiries.", "Tutor changes marketing focus.", "V7", 5),
    task("DATA-021", "P2", "DATA", "Monthly business review", "Create monthly review of inquiries, trials, fees, proof, source quality, and next recommendations.", "Data/Product", "P2", "Not Started", date(2027, 7, 1), date(2027, 8, 15), "DATA-020,FE-042", "Review gives concrete next actions.", "Owners review monthly.", "V7", 6),
    task("STR-020", "P2", "STR", "Scale segment decision", "Decide whether to expand to exam coaches, hobby teachers, or double down on academic tutors.", "Founder/Product", "P1", "Not Started", date(2027, 8, 15), date(2027, 9, 12), "DATA-021,GTM-011", "Decision memo uses retention, revenue, support, demand evidence.", "Next segment chosen with evidence.", "V7", 3),
    task("AI-010", "P2", "AI", "Learner Context Card V1", "Add context card: class, board, concern, confidence, language comfort, parent goal, next tutor action.", "AI/Product/Frontend", "P2", "Not Started", date(2027, 9, 13), date(2027, 11, 1), "FE-042,STR-020", "Context card is editable, gentle, and non-judgmental.", "Tutors use it in parent conversations.", "V8", 8),
    task("SEC-020", "P2", "SEC", "Learning analytics safety policy", "Define consent, correction, deletion, non-labeling, visibility, confidence and human-confirmation rules.", "Security/Product", "P2", "Not Started", date(2027, 9, 13), date(2027, 10, 15), "AI-010", "Safety policy governs all learner analytics.", "No harmful fixed labels.", "V8", 5),
    task("AI-011", "P2", "AI", "Intervention action recommendations", "Suggest tutor-confirmed next actions based on attendance, concerns, goals, and progress notes.", "AI/Product", "P2", "Not Started", date(2027, 11, 1), date(2027, 12, 31), "AI-010,SEC-020", "Recommendations are explainable and editable.", "Interventions completed.", "V8", 7),
    task("AI-012", "P2", "AI", "Practical learning prompts", "Generate real-world examples and small projects linked to subject, learner context, local life, planet, and careers.", "AI/Product", "P2", "Not Started", date(2028, 1, 1), date(2028, 3, 12), "AI-011", "Prompts connect concepts to practical life.", "Tutors use prompts in lessons.", "V9", 7),
    task("INST-001", "P2", "INST", "Institution ICP research", "Research schools/colleges/institutes needing learner support analytics and parent communication intelligence.", "Founder/Product", "P2", "Not Started", date(2028, 3, 13), date(2028, 6, 30), "AI-012", "Institution pain map and buying process documented.", "3 institution discovery calls completed.", "V10", 5),
    task("INST-002", "P2", "INST", "Multi-teacher account model", "Add roles for owner, teacher, coordinator, admin, support and cohort-level visibility.", "Product/Backend/Security", "P2", "Not Started", date(2028, 7, 1), date(2028, 10, 31), "INST-001,D-002", "Role model supports institution pilot.", "No cross-role data leakage.", "V10", 8),
    task("INST-003", "P2", "INST", "Cohort dashboard pilot", "Build cohort progress, attendance, intervention and parent communication dashboard.", "Product/Data/Frontend", "P2", "Not Started", date(2028, 11, 1), date(2029, 3, 31), "INST-002,AI-011", "Dashboard supports one institution pilot without ERP replacement.", "Pilot admin uses weekly.", "V10", 9),
    task("INST-004", "P2", "INST", "Institution pilot launch", "Launch controlled pilot with school, college, or coaching institution.", "Founder/Product/GTM", "P2", "Not Started", date(2029, 4, 1), date(2029, 6, 30), "INST-003", "Pilot contract, success metrics, support plan, data policy signed.", "Institution pilot active.", "V10", 8),
]


VISION = [
    ["V1", "Tutor Trust Entry", "Tutors must look credible before parents inquire.", "Public page, proof, FAQ, local context.", "Public Trust Page", "Tutor says they would share page."],
    ["V2", "Structured Context Capture", "Raw WhatsApp lacks class, board, concern, timing context.", "Inquiry form captures minimum useful learning/business context.", "Structured Inquiry", "Owner can reply without basic questions."],
    ["V3", "WhatsApp-Native Operations", "Tutors will not abandon WhatsApp.", "Editable click-to-chat templates and action statuses.", "WhatsApp Handoff", "Owner sends/records follow-ups weekly."],
    ["V4", "Trial To Enrollment Discipline", "Missed demos leak revenue.", "Trial status, reminders, post-trial follow-up, enrolled/lost reasons.", "Trial Workflow", "Trial outcomes are marked."],
    ["V5", "Fee Clarity Without Awkwardness", "Recurring cash flow is relationship-sensitive.", "Fee cycles, ageing, polite reminders, mark paid/part paid.", "Fee Cycle", "Tutor tracks at least one real cycle."],
    ["V6", "Proof And Referral Loop", "Good teaching often stays invisible.", "Testimonials, result proof, referral requests, approval workflow.", "Proof Loop", "Tutor collects proof monthly."],
    ["V7", "Owner Operating Rhythm", "Dashboard must produce action, not decoration.", "Today queue, weekly digest, readiness and source signals.", "Operating Rhythm", "Weekly completed tutor actions increase."],
    ["V8", "Learner Context", "The larger vision requires understanding learners beyond marks.", "Learner context card, parent goals, confidence, concerns, next action.", "Learner Context Card", "Tutors use context in parent conversations."],
    ["V9", "Practical Learning", "Learning must connect to life, work, planet, community.", "Project prompts, real-world examples, contextual lesson enrichment.", "Practical Learning Engine", "Assignments include real-world context."],
    ["V10", "Institutional Intelligence", "Tutors are entry point to schools/colleges.", "Cohort dashboards, intervention tracking, improvement reports.", "Institution Dashboards", "Institution pilot signed."],
]

RISKS = [
    ["R-001", "Identity drift between BizzNexx and ContexureEarth", "High", "High", "Confuses customers and roadmap.", "Make ContexureEarth public education brand; BizzNexx internal/legacy engine.", "Founder", "Open", "IMD-006"],
    ["R-002", "Overbuilding LMS/ERP features", "High", "Medium", "Slows launch and weakens wedge.", "Use decision rules and P3 deferral list.", "Product", "Open", "STR-002"],
    ["R-003", "Tutor expects guaranteed leads", "High", "Medium", "Creates mistrust and churn.", "Position as conversion/operations, not marketplace.", "GTM", "Open", "GTM-001"],
    ["R-004", "Parent/student data leakage", "Critical", "Low", "Trust and legal risk.", "Tenant isolation, auth, consent, deletion, minimization.", "Security", "Open", "D-002"],
    ["R-005", "WhatsApp automation too early", "Medium", "Medium", "Compliance and complexity risk.", "Launch manual handoff first.", "Product", "Open", "F-001"],
    ["R-006", "Fee reminders feel aggressive", "Medium", "Medium", "Relationship damage.", "Polite templates and manual review before send.", "UX", "Open", "F-003"],
    ["R-007", "Weak weekly retention", "High", "Medium", "Product becomes setup service only.", "Build owner action rhythm and weekly digest.", "Data/Product", "Open", "DATA-002"],
    ["R-008", "AI overclaims learner ability", "Critical", "Medium", "Can harm learners and trust.", "Human-confirmed, editable, non-judgmental insights.", "AI/Security", "Open", "SEC-020"],
    ["R-009", "Support burden too high", "Medium", "Medium", "Margins suffer.", "Setup SOP, help center, fewer custom workflows.", "OPS", "Open", "OPS-001"],
    ["R-010", "No paid willingness", "High", "Medium", "Business model not validated.", "Paid pilot pricing tests and cancellation interviews.", "GTM", "Open", "GTM-004"],
]

GATES = [
    ["Alpha Gate", "P1", date(2026, 7, 19), "Public page works; inquiry flow works; owner queue exists; WhatsApp handoff works; fee tracker exists; demo data labeled.", "C-003,C-004,C-005,C-008,F-001", "Not Started", ""],
    ["Private Beta Gate", "P2", date(2026, 9, 6), "Owner login; tenant isolation; persistence; consent; beta terms; backup plan; smoke tests; mobile QA; support channel.", "SEC-010,D-002,E-006,SEC-014,INFRA-010,QA-011", "Not Started", ""],
    ["Paid Pilot Gate", "P3", date(2026, 12, 13), "Repeatable setup; pricing clear; support docs; source tracking; weekly digest; proof/referral; case template.", "GTM-001,OPS-001,DATA-002,FE-020,FE-021,GTM-005", "Not Started", ""],
    ["Public Launch Gate", "P4", date(2027, 3, 14), "Production monitoring; rollback; privacy; deletion flow; onboarding; billing; conversion dashboard; churn tracking.", "INFRA-020,FE-030,GTM-010,DATA-010,QA-020", "Not Started", ""],
    ["Intelligence Gate", "P5", date(2027, 9, 12), "AI drafts accepted; analytics useful; support load manageable; data safety stable.", "AI-001,AI-002,DATA-021,STR-020", "Not Started", ""],
    ["Learning Context Gate", "P6", date(2028, 3, 12), "Learner context used ethically; parent reports constructive; consent/correction ready.", "AI-010,SEC-020,AI-011,AI-012", "Not Started", ""],
    ["Institution Gate", "P7", date(2029, 6, 30), "Tutor product retained; data privacy mature; institution pilot has signed success criteria.", "INST-001,INST-002,INST-003,INST-004", "Not Started", ""],
]

METRICS = [
    ["North Star", "Weekly completed tutor actions", "Count inquiry replies, trial follow-ups, fee updates, reminders, proof/referral/progress actions completed weekly.", "Weekly", "Product/Data", "P2", "25+ actions/week across beta by private beta end"],
    ["Activation", "Setup completed", "Tutor profile, courses, proof, FAQ, templates and public page ready.", "Weekly", "OPS", "P1", "70% of onboarded tutors complete setup"],
    ["Activation", "Public page shared", "Tutor shares link with parent/prospect.", "Weekly", "GTM/Data", "P1", "3 of first 5 pilot tutors share page"],
    ["Activation", "First inquiry captured", "At least one real inquiry submitted.", "Weekly", "Product/Data", "P1", "20 real inquiries by concierge exit"],
    ["Retention", "Dashboard weekly active tutors", "Tutors visiting action queue weekly.", "Weekly", "Data", "P2", "50% private beta WAU"],
    ["Workflow", "Inquiry status update rate", "Share of inquiries with status updated by owner.", "Weekly", "Data", "P2", "30%+ in private beta"],
    ["Revenue", "Paid pilot conversion", "Beta users converting to paid pilot.", "Monthly", "GTM", "P3", "10 paid pilots"],
    ["Revenue", "Month-two retention", "Paid pilots active in second month.", "Monthly", "GTM/Data", "P3", "60%+"],
    ["Quality", "Setup time", "Time to create share-ready tutor account.", "Monthly", "OPS", "P3", "Under 90 minutes"],
    ["Learning", "Learner context card usage", "Learners with useful context card reviewed by tutor.", "Monthly", "AI/Data", "P6", "50% of active learners in context pilots"],
]

BACKLOG = [
    ["BKL-001", "Google Forms import", "P2", "Scale", "Useful for tutors already using Forms.", "After native inquiry validates.", ""],
    ["BKL-002", "Google Sheets export", "P2", "Scale", "Useful for trust and data portability.", "After paid pilot.", ""],
    ["BKL-003", "Payment gateway", "P3", "Deferred", "Adds complexity before manual fee value proven.", "After fee tracker retention.", ""],
    ["BKL-004", "Parent mobile app", "P3", "Deferred", "Too much scope before owner workflow retention.", "After institution/tutor demand.", ""],
    ["BKL-005", "Full LMS", "P3", "Deferred", "Competes with LMS and dilutes wedge.", "Only if paid users demand.", ""],
    ["BKL-006", "WhatsApp Cloud API automation", "P2/P3", "Scale", "Useful after manual workflow proves value and opt-in is clear.", "After 100+ active tutors.", ""],
    ["BKL-007", "Corporate trainer proposal workflow", "P3", "Deferred", "Different ICP and sales cycle.", "After tutor engine matures.", ""],
    ["BKL-008", "Full test engine", "P3", "Deferred", "Exam-coach specific and data-heavy.", "After exam coach segment validated.", ""],
]


def style_sheet(ws, freeze=True):
    ws.freeze_panes = "A2" if freeze else None
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E5E7EB")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"


def set_widths(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = width


def add_table(ws, name):
    max_row = ws.max_row
    max_col = ws.max_column
    end_col = ws.cell(1, max_col).column_letter
    ref = f"A1:{end_col}{max_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)


def add_validations(ws):
    dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUS) + '"', allow_blank=False)
    dv_pri = DataValidation(type="list", formula1='"' + ",".join(PRIORITY) + '"', allow_blank=False)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_pri)
    dv_status.add(f"H2:H{ws.max_row}")
    dv_pri.add(f"G2:G{ws.max_row}")


def make_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # Dashboard
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ContexureEarth Tutor Project Command Center"
    ws["A1"].font = Font(size=18, bold=True, color="0F172A")
    ws.merge_cells("A1:G1")
    ws["A2"] = "Purpose"
    ws["B2"] = "Vision-linked project planner covering what is already done, what remains, dependencies, launch gates, risks, research, and metrics."
    ws.merge_cells("B2:G2")
    ws["A3"] = "North Star"
    ws["B3"] = "Weekly completed tutor actions"
    ws["A4"] = "Last Updated"
    ws["B4"] = date(2026, 6, 13)
    ws["B4"].number_format = "yyyy-mm-dd"
    dashboard_rows = [
        ("Total Tasks", '=COUNTA(\'Master Plan\'!A2:A200)', "All tracked tasks"),
        ("Done Tasks", '=COUNTIF(\'Master Plan\'!H2:H200,"Done")', "Completed BizzNexx/ContexureEarth work"),
        ("Open P0 Tasks", '=COUNTIFS(\'Master Plan\'!G2:G200,"P0",\'Master Plan\'!H2:H200,"<>Done")', "Critical work left"),
        ("Blocked Tasks", '=COUNTIF(\'Master Plan\'!H2:H200,"Blocked")', "Needs founder/engineering attention"),
        ("Open Critical Risks", '=COUNTIFS(\'Risk Register\'!D2:D100,"Critical",\'Risk Register\'!H2:H100,"Open")', "No launch with critical open risk"),
        ("Private Beta Gate Date", date(2026, 9, 6), "Current target gate"),
    ]
    ws.append([])
    ws.append(["KPI", "Value", "Meaning"])
    for r in dashboard_rows:
        ws.append(r)
    ws.append([])
    ws.append(["Phase", "Start", "End", "Mission"])
    for _, phase, start, end in PHASES:
        ws.append([phase, start, end, next((x[3] for x in GATES if x[1] in phase), "")])
    style_sheet(ws, freeze=False)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 70
    for row in range(6, 13):
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="E0F2FE")
        ws[f"B{row}"].font = Font(bold=True, color="0F172A")

    # Master Plan
    master = wb.create_sheet("Master Plan")
    headers = [
        "Task ID", "Phase", "Workstream", "Title", "Description", "Owner", "Priority", "Status", "Start",
        "End", "Dependencies", "Effort", "Acceptance Criteria", "Validation Signal", "Vision Link", "Notes",
    ]
    master.append(headers)
    for row in DONE_TASKS + FUTURE_TASKS:
        master.append(row)
    style_sheet(master)
    widths = [14, 10, 26, 34, 62, 22, 10, 16, 13, 13, 30, 10, 66, 50, 18, 34]
    for col, width in enumerate(widths, start=1):
        master.column_dimensions[master.cell(1, col).column_letter].width = width
    add_validations(master)
    add_table(master, "MasterPlanTable")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    amber_fill = PatternFill("solid", fgColor="FEF3C7")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    master.conditional_formatting.add(f"H2:H{master.max_row}", FormulaRule(formula=['$H2="Blocked"'], fill=red_fill))
    master.conditional_formatting.add(f"H2:H{master.max_row}", FormulaRule(formula=['$H2="In Progress"'], fill=amber_fill))
    master.conditional_formatting.add(f"H2:H{master.max_row}", FormulaRule(formula=['$H2="Done"'], fill=green_fill))

    # Done Inventory
    done = wb.create_sheet("Done Inventory")
    done.append(headers)
    for row in DONE_TASKS:
        done.append(row)
    style_sheet(done)
    for col, width in enumerate(widths, start=1):
        done.column_dimensions[done.cell(1, col).column_letter].width = width
    add_table(done, "DoneInventoryTable")

    # Remaining Work
    rem = wb.create_sheet("Remaining Work")
    rem.append(headers)
    for row in FUTURE_TASKS:
        rem.append(row)
    style_sheet(rem)
    for col, width in enumerate(widths, start=1):
        rem.column_dimensions[rem.cell(1, col).column_letter].width = width
    add_validations(rem)
    add_table(rem, "RemainingWorkTable")

    # Vision map
    vis = wb.create_sheet("Vision Map")
    vis.append(["Vision ID", "Pillar", "Founder Belief", "Product Translation", "Product Surface", "Validation Signal"])
    for row in VISION:
        vis.append(row)
    style_sheet(vis)
    for col, width in enumerate([12, 28, 56, 58, 30, 46], start=1):
        vis.column_dimensions[vis.cell(1, col).column_letter].width = width
    add_table(vis, "VisionMapTable")

    # Timeline
    tl = wb.create_sheet("Timeline")
    tl.append(["Phase ID", "Phase", "Start", "End", "Duration Days", "Target Gate"])
    for pid, phase, start, end in PHASES:
        tl.append([pid, phase, start, end, f"=D{tl.max_row+1}-C{tl.max_row+1}+1", next((g[0] for g in GATES if g[1] == pid), "")])
    style_sheet(tl)
    for col, width in enumerate([12, 42, 14, 14, 18, 28], start=1):
        tl.column_dimensions[tl.cell(1, col).column_letter].width = width
    add_table(tl, "TimelineTable")

    # Dependencies
    dep = wb.create_sheet("Dependencies")
    dep.append(["Task ID", "Depends On", "Dependency Type", "Why It Matters", "Blocked If Missing"])
    for row in DONE_TASKS + FUTURE_TASKS:
        if row[10]:
            for d in str(row[10]).split(","):
                dep.append([row[0], d.strip(), "Finish-to-start", "Task relies on upstream output being complete.", "Yes"])
    style_sheet(dep)
    for col, width in enumerate([16, 18, 18, 58, 20], start=1):
        dep.column_dimensions[dep.cell(1, col).column_letter].width = width
    add_table(dep, "DependencyTable")

    # Gates
    gates = wb.create_sheet("Launch Gates")
    gates.append(["Gate", "Phase", "Target Date", "Criteria", "Required Tasks", "Status", "Evidence / Notes"])
    for row in GATES:
        gates.append(row)
    style_sheet(gates)
    for col, width in enumerate([24, 10, 14, 78, 58, 16, 38], start=1):
        gates.column_dimensions[gates.cell(1, col).column_letter].width = width
    add_table(gates, "LaunchGatesTable")

    # Risks
    risk = wb.create_sheet("Risk Register")
    risk.append(["Risk ID", "Risk", "Impact", "Likelihood", "Why It Matters", "Mitigation", "Owner", "Status", "Linked Task"])
    for row in RISKS:
        risk.append(row)
    style_sheet(risk)
    for col, width in enumerate([12, 42, 14, 16, 54, 62, 22, 14, 16], start=1):
        risk.column_dimensions[risk.cell(1, col).column_letter].width = width
    add_table(risk, "RiskRegisterTable")

    # Metrics
    metrics = wb.create_sheet("Metrics")
    metrics.append(["Metric Type", "Metric", "Definition", "Cadence", "Owner", "First Relevant Phase", "Target"])
    for row in METRICS:
        metrics.append(row)
    style_sheet(metrics)
    for col, width in enumerate([18, 32, 76, 16, 22, 20, 46], start=1):
        metrics.column_dimensions[metrics.cell(1, col).column_letter].width = width
    add_table(metrics, "MetricsTable")

    # Research tracker
    research = wb.create_sheet("Research Tracker")
    research.append(["Research ID", "Name", "Segment", "City", "Contact", "Status", "Scheduled Date", "Completed Date", "Current Tools", "Top Pain", "Pain Score 1-5", "WTP Setup", "WTP Monthly", "Key Quote", "Next Step"])
    for i in range(1, 41):
        research.append([f"RCH-{i:03d}", "", "", "", "", "Not Contacted", "", "", "", "", "", "", "", "", ""])
    style_sheet(research)
    for col, width in enumerate([14, 24, 20, 16, 20, 16, 15, 15, 30, 44, 16, 14, 14, 54, 30], start=1):
        research.column_dimensions[research.cell(1, col).column_letter].width = width
    add_table(research, "ResearchTrackerTable")

    # Backlog
    backlog = wb.create_sheet("Backlog")
    backlog.append(["Backlog ID", "Item", "Priority", "Category", "Why It Exists", "Trigger To Build", "Notes"])
    for row in BACKLOG:
        backlog.append(row)
    style_sheet(backlog)
    for col, width in enumerate([14, 38, 12, 18, 58, 48, 36], start=1):
        backlog.column_dimensions[backlog.cell(1, col).column_letter].width = width
    add_table(backlog, "BacklogTable")

    # Dashboard charts from helper data
    summary = wb.create_sheet("Summary Data")
    summary.append(["Status", "Count"])
    for status in STATUS:
        summary.append([status, f'=COUNTIF(\'Master Plan\'!H:H,"{status}")'])
    summary.append([])
    summary.append(["Priority", "Open Count"])
    for p in PRIORITY:
        summary.append([p, f'=COUNTIFS(\'Master Plan\'!G:G,"{p}",\'Master Plan\'!H:H,"<>Done")'])
    style_sheet(summary)
    summary.sheet_state = "hidden"

    status_chart = PieChart()
    labels = Reference(summary, min_col=1, min_row=2, max_row=8)
    data = Reference(summary, min_col=2, min_row=1, max_row=8)
    status_chart.add_data(data, titles_from_data=True)
    status_chart.set_categories(labels)
    status_chart.title = "Task Status Mix"
    ws.add_chart(status_chart, "E6")

    pri_chart = BarChart()
    pri_labels = Reference(summary, min_col=1, min_row=11, max_row=14)
    pri_data = Reference(summary, min_col=2, min_row=10, max_row=14)
    pri_chart.add_data(pri_data, titles_from_data=True)
    pri_chart.set_categories(pri_labels)
    pri_chart.title = "Open Tasks By Priority"
    pri_chart.y_axis.title = "Open Tasks"
    pri_chart.x_axis.title = "Priority"
    ws.add_chart(pri_chart, "E22")

    # Lists
    lists = wb.create_sheet("Lists")
    lists.append(["Statuses", "Priorities", "Phases", "Workstreams"])
    max_len = max(len(STATUS), len(PRIORITY), len(PHASES), len(WORKSTREAMS))
    for i in range(max_len):
        lists.append([
            STATUS[i] if i < len(STATUS) else "",
            PRIORITY[i] if i < len(PRIORITY) else "",
            PHASES[i][0] if i < len(PHASES) else "",
            list(WORKSTREAMS.values())[i] if i < len(WORKSTREAMS) else "",
        ])
    style_sheet(lists)
    lists.sheet_state = "hidden"

    wb.properties.title = "ContexureEarth Tutor Professional Project Planner"
    wb.properties.creator = "Codex"
    return wb


if __name__ == "__main__":
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = make_workbook()
    wb.save(OUT)
    # Verification pass: reopen and assert expected sheets/content exists.
    check = load_workbook(OUT, data_only=False)
    required = {"Dashboard", "Master Plan", "Done Inventory", "Remaining Work", "Vision Map", "Timeline", "Dependencies", "Launch Gates", "Risk Register", "Metrics", "Research Tracker", "Backlog"}
    missing = required - set(check.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {missing}")
    if check["Master Plan"].max_row < 80:
        raise RuntimeError("Master Plan has too few rows")
    print(f"Created {OUT}")
    print(f"Master Plan rows: {check['Master Plan'].max_row - 1}")
    print(f"Done rows: {check['Done Inventory'].max_row - 1}")
    print(f"Remaining rows: {check['Remaining Work'].max_row - 1}")
